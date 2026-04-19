"""
ILM1300 Spreadsheet Importer
Reads PartList_private.xlsx and ANH_donors.xlsx into the SQLite database.

Usage:
    python import_spreadsheets.py --kits    path/to/PartList_private.xlsx
    python import_spreadsheets.py --donors  path/to/ANH_donors.xlsx
    python import_spreadsheets.py --kits path/to/PartList_private.xlsx \
                                  --donors path/to/ANH_donors.xlsx
"""

import sqlite3
import os
import sys
import argparse
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, 'data', 'ilm1300.db')


# ── DB HELPERS ────────────────────────────────────────────────────

def get_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON")
    return db

def clean(val):
    """Strip whitespace and return None for empty/None/placeholder values."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ('none', 'nan', '—', '-') else None

def safe_int(val):
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None

def safe_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


# ── ENSURE FALCON MODEL EXISTS ────────────────────────────────────

def ensure_falcon(db):
    existing = db.execute("SELECT id FROM models WHERE slug='falcon-5ft'").fetchone()
    if existing:
        return existing['id']
    cur = db.execute("""
        INSERT INTO models (name, slug, film, scale_approx, notes)
        VALUES ('Millennium Falcon (5ft)', 'falcon-5ft', 'ANH',
                '1/20.5', 'Original 5-foot hero Falcon built at ILM Van Nuys')
    """)
    db.commit()
    print("  Created model: Millennium Falcon (5ft)")
    return cur.lastrowid


# ── IMPORT KITS SHEET ─────────────────────────────────────────────

def import_kits(xlsx_path, db):
    print(f"\nImporting kits from: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if 'kits' not in wb.sheetnames:
        print("  ERROR: No 'kits' sheet found")
        return 0

    ws   = wb['kits']
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (has 'brand' and 'name')
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'brand' in vals and 'name' in vals:
            header_row = i
            break

    if header_row is None:
        print("  ERROR: Could not find header row in kits sheet")
        return 0

    headers = [str(v).strip().lower() if v else f'col{i}'
               for i, v in enumerate(rows[header_row])]
    print(f"  Headers: {[h for h in headers if h and not h.startswith('col')]}")

    def col(row, name):
        try:
            idx = headers.index(name)
            return row[idx] if idx < len(row) else None
        except ValueError:
            return None

    imported = 0
    skipped  = 0

    for row in rows[header_row + 1:]:
        brand = clean(col(row, 'brand'))
        name  = clean(col(row, 'name'))

        if not brand or not name:
            continue

        # Parse Coffman reference values (stored in kit_references, not kits)
        num_raw  = safe_float(col(row, '#'))
        base_raw = safe_float(col(row, 'base'))
        suff_raw = safe_float(col(row, 'suffix'))

        coffman_num    = safe_int(num_raw)    if num_raw    is not None else None
        coffman_base   = safe_int(base_raw)   if base_raw   is not None else None
        coffman_suffix = safe_int(suff_raw)   if suff_raw   is not None else None

        scale          = clean(col(row, 'scale'))
        serial_number  = clean(col(row, 'sn'))
        scalemates_url = clean(col(row, 'scalemates'))
        scans_url      = clean(col(row, 'kit scans'))
        instructions   = clean(col(row, 'instructions'))
        notes          = clean(col(row, 'additional info'))

        # Skip if already exists (brand + name + serial)
        existing = db.execute("""
            SELECT id FROM kits WHERE brand=? AND name=?
            AND (serial_number=? OR (serial_number IS NULL AND ? IS NULL))
        """, (brand, name, serial_number, serial_number)).fetchone()

        if existing:
            kit_id = existing['id']
            skipped += 1
        else:
            cur = db.execute("""
                INSERT INTO kits
                    (brand, scale, name, serial_number,
                     scalemates_url, scans_url, instructions_url, notes)
                VALUES (?,?,?,?,?,?,?,?)
            """, (brand, scale, name, serial_number,
                  scalemates_url, scans_url, instructions, notes))
            kit_id = cur.lastrowid
            imported += 1

        # Store Coffman reference in kit_references (system='coffman')
        # Use the full number as value if available, otherwise base+suffix
        if coffman_num is not None:
            coffman_value = str(coffman_num) if coffman_suffix == 0 \
                            else f"{coffman_base}{coffman_suffix}"
            db.execute("""
                INSERT OR IGNORE INTO kit_references (kit_id, system, value)
                VALUES (?, 'coffman', ?)
            """, (kit_id, coffman_value))

    db.commit()
    print(f"  Kits imported: {imported}  |  skipped (already exist): {skipped}")
    wb.close()
    return imported


# ── IMPORT PARTS SHEET ────────────────────────────────────────────

def import_parts(xlsx_path, db):
    print(f"\nImporting parts from: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if 'parts' not in wb.sheetnames:
        print("  No 'parts' sheet found — skipping")
        return 0

    ws   = wb['parts']
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (has 'kit #-base' or 'part id')
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'kit #-base' in vals or 'part id' in vals:
            header_row = i
            break

    if header_row is None:
        print("  Could not find header row in parts sheet — skipping")
        return 0

    headers = [str(v).strip().lower() if v else f'col{i}'
               for i, v in enumerate(rows[header_row])]

    def col(row, name):
        try:
            idx = headers.index(name)
            return row[idx] if idx < len(row) else None
        except ValueError:
            return None

    falcon_id = ensure_falcon(db)
    map_cache = {}  # map_name -> map_id

    imported = 0
    skipped  = 0
    no_kit   = 0

    for row in rows[header_row + 1:]:
        kit_base = safe_int(safe_float(col(row, 'kit #-base')))
        part_num = clean(col(row, 'part id'))
        map_name = clean(col(row, 'map / plate'))

        if kit_base is None or part_num is None:
            continue

        # Look up kit via kit_references table (system='coffman')
        # This replaces the old direct coffman_base column lookup
        ref = db.execute("""
            SELECT kr.kit_id FROM kit_references kr
            WHERE kr.system = 'coffman' AND kr.value = ?
            LIMIT 1
        """, (str(kit_base),)).fetchone()

        if not ref:
            no_kit += 1
            continue

        kit_id = ref['kit_id']

        # Ensure part exists
        existing_part = db.execute(
            "SELECT id FROM parts WHERE kit_id=? AND part_number=?",
            (kit_id, part_num)
        ).fetchone()

        if existing_part:
            part_id = existing_part['id']
            skipped += 1
        else:
            notes_raw = clean(col(row, 'col7'))
            cur = db.execute("""
                INSERT INTO parts (kit_id, part_number, notes)
                VALUES (?,?,?)
            """, (kit_id, part_num, notes_raw))
            part_id  = cur.lastrowid
            imported += 1

        # Ensure map exists and create placement
        if map_name:
            if map_name not in map_cache:
                existing_map = db.execute(
                    "SELECT id FROM maps WHERE model_id=? AND name=?",
                    (falcon_id, map_name)
                ).fetchone()
                if existing_map:
                    map_cache[map_name] = existing_map['id']
                else:
                    cur = db.execute(
                        "INSERT INTO maps (model_id, name) VALUES (?,?)",
                        (falcon_id, map_name)
                    )
                    map_cache[map_name] = cur.lastrowid

            map_id     = map_cache[map_name]
            copy_count = safe_int(safe_float(col(row, 'copys on map'))) or 1

            # Avoid duplicate placements
            existing_pl = db.execute("""
                SELECT id FROM placements
                WHERE model_id=? AND map_id=? AND part_id=?
            """, (falcon_id, map_id, part_id)).fetchone()

            if not existing_pl:
                notes_raw = clean(col(row, 'col7'))
                db.execute("""
                    INSERT INTO placements
                        (model_id, map_id, part_id, copy_count, notes)
                    VALUES (?,?,?,?,?)
                """, (falcon_id, map_id, part_id, copy_count, notes_raw))

    db.commit()
    print(f"  Parts imported: {imported}  |  already existed: {skipped}  |  kit not found: {no_kit}")
    print(f"  Maps created/used: {len(map_cache)}")
    wb.close()
    return imported


# ── IMPORT MAPS SHEET ─────────────────────────────────────────────

def import_maps(xlsx_path, db):
    print(f"\nImporting maps from: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if 'maps' not in wb.sheetnames:
        print("  No 'maps' sheet — skipping")
        return 0

    ws        = wb['maps']
    falcon_id = ensure_falcon(db)
    imported  = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name    = clean(row[0])
        url     = clean(row[1])
        version = clean(row[2])
        if not name:
            continue

        existing = db.execute(
            "SELECT id FROM maps WHERE model_id=? AND name=?",
            (falcon_id, name)
        ).fetchone()

        if existing:
            if url or version:
                db.execute("""
                    UPDATE maps SET url=COALESCE(?,url), version=COALESCE(?,version)
                    WHERE id=?
                """, (url, version, existing['id']))
        else:
            db.execute("""
                INSERT INTO maps (model_id, name, url, version)
                VALUES (?,?,?,?)
            """, (falcon_id, name, url, version))
            imported += 1

    db.commit()
    print(f"  Maps imported/updated: {imported}")
    wb.close()
    return imported


# ── IMPORT 3D PARTS SHEET ─────────────────────────────────────────

def import_3d_parts(xlsx_path, db):
    """
    Imports the 'existing 3D parts' sheet into part_files.
    Sheet columns: kit (coffman#), part, provider 1, provider 2, provider 3, notes
    """
    print(f"\nImporting 3D parts from: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    sheet_name = 'existing 3D parts'
    if sheet_name not in wb.sheetnames:
        print(f"  No '{sheet_name}' sheet — skipping")
        return 0

    ws       = wb[sheet_name]
    rows     = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    imported = 0
    no_kit   = 0
    no_part  = 0

    for row in rows:
        coffman_num = safe_int(safe_float(row[0] if len(row) > 0 else None))
        part_num    = clean(row[1] if len(row) > 1 else None)
        providers   = [clean(row[i]) for i in range(2, 5) if i < len(row)]
        notes       = clean(row[5] if len(row) > 5 else None)

        if coffman_num is None or part_num is None:
            continue

        # Look up kit via kit_references
        ref = db.execute("""
            SELECT kr.kit_id FROM kit_references kr
            WHERE kr.system = 'coffman' AND kr.value = ?
            LIMIT 1
        """, (str(coffman_num),)).fetchone()

        if not ref:
            no_kit += 1
            continue

        kit_id = ref['kit_id']

        # Look up part
        part = db.execute(
            "SELECT id FROM parts WHERE kit_id=? AND part_number=?",
            (kit_id, part_num)
        ).fetchone()

        if not part:
            no_part += 1
            continue

        part_id = part['id']

        # Create a part_files entry per provider that has a value
        for provider in providers:
            if not provider:
                continue
            existing = db.execute("""
                SELECT id FROM part_files
                WHERE part_id=? AND source=? AND file_type='scan'
            """, (part_id, provider)).fetchone()

            if not existing:
                db.execute("""
                    INSERT INTO part_files (part_id, file_type, url, source, notes)
                    VALUES (?, 'scan', NULL, ?, ?)
                """, (part_id, provider, notes))
                imported += 1

    db.commit()
    print(f"  3D part entries imported: {imported}  |  kit not found: {no_kit}  |  part not found: {no_part}")
    wb.close()
    return imported


# ── IMPORT ANH DONORS (CROSS-MODEL) ──────────────────────────────

def import_donors(xlsx_path, db):
    print(f"\nImporting ANH donors from: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if 'SWANH donors' not in wb.sheetnames:
        print("  No 'SWANH donors' sheet found")
        return 0

    ws   = wb['SWANH donors']
    rows = list(ws.iter_rows(values_only=True))

    model_row  = rows[1]

    MODEL_SLUGS = {
        'Millennium Falcon':      'falcon-5ft',
        'X-Wing':                 'x-wing',
        'Y-Wing':                 'y-wing',
        'Tie Fighter':            'tie-fighter',
        'Tie Advanced':           'tie-advanced',
        'Star Destroyer':         'star-destroyer',
        'Death Star Turbo Laser': 'death-star-turbolaser',
        'Death Star Tiles 1-6':   'death-star-tiles',
        'Training Remote':        'training-remote',
        'Blockade Runner':        'blockade-runner',
        'Sandcrawler':            'sandcrawler',
        'Escape Pod':             'escape-pod',
        'Death Star Crane':       'death-star-crane',
    }

    MODEL_COLS = {}
    for i, val in enumerate(model_row):
        name = clean(val)
        if name and name in MODEL_SLUGS:
            MODEL_COLS[i] = MODEL_SLUGS[name]

    # Ensure all models exist
    for name, slug in MODEL_SLUGS.items():
        existing = db.execute("SELECT id FROM models WHERE slug=?", (slug,)).fetchone()
        if not existing:
            db.execute("""
                INSERT INTO models (name, slug, film) VALUES (?,?,'ANH')
            """, (name, slug))
    db.commit()

    imported = 0
    for row in rows[2:]:
        brand  = clean(row[0])
        scale  = clean(row[1])
        name   = clean(row[2])
        sn     = clean(row[3])
        sm_url = clean(row[4])
        scans  = clean(row[5])

        if not brand or not name:
            continue

        # Find or create kit
        kit = db.execute(
            "SELECT id FROM kits WHERE brand=? AND name=?", (brand, name)
        ).fetchone()

        if not kit:
            cur = db.execute("""
                INSERT INTO kits (brand, scale, name, serial_number, scalemates_url, scans_url)
                VALUES (?,?,?,?,?,?)
            """, (brand, scale, name, sn, sm_url, scans))
            kit_id = cur.lastrowid
        else:
            kit_id = kit['id']

        # Record which models this kit appears on.
        # No part-level detail from this sheet — use kit_id directly on placements
        # (the new schema allows this; no '?' placeholder parts needed)
        for col_idx, slug in MODEL_COLS.items():
            if col_idx >= len(row):
                continue
            val = clean(row[col_idx])
            if not val:
                continue

            model = db.execute("SELECT id FROM models WHERE slug=?", (slug,)).fetchone()
            if not model:
                continue

            # Avoid duplicate kit-level placements
            existing = db.execute("""
                SELECT id FROM placements
                WHERE model_id=? AND kit_id=?
                AND part_id IS NULL AND cast_assembly_id IS NULL
            """, (model['id'], kit_id)).fetchone()

            if not existing:
                confidence = 'confirmed' if val.lower() == 'x' else 'probable'
                db.execute("""
                    INSERT INTO placements (model_id, kit_id, confidence, notes)
                    VALUES (?,?,?,?)
                """, (model['id'], kit_id, confidence,
                      val if val.lower() != 'x' else None))
                imported += 1

    db.commit()
    print(f"  Cross-model placements recorded: {imported}")
    wb.close()
    return imported


# ── MAIN ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Import spreadsheets into ILM1300 database')
    parser.add_argument('--kits',   help='Path to PartList_private.xlsx')
    parser.add_argument('--donors', help='Path to ANH_donors.xlsx')
    args = parser.parse_args()

    if not args.kits and not args.donors:
        parser.print_help()
        sys.exit(1)

    if not os.path.exists(DB_PATH):
        print(f"ERROR: Database not found at {DB_PATH}")
        print("Run 'python app.py' once first to initialise the database.")
        sys.exit(1)

    db = get_db()

    if args.kits:
        if not os.path.exists(args.kits):
            print(f"ERROR: File not found: {args.kits}")
            sys.exit(1)
        import_kits(args.kits, db)
        import_maps(args.kits, db)
        import_parts(args.kits, db)
        import_3d_parts(args.kits, db)

    if args.donors:
        if not os.path.exists(args.donors):
            print(f"ERROR: File not found: {args.donors}")
            sys.exit(1)
        import_donors(args.donors, db)

    db.close()
    print("\nImport complete.")

if __name__ == '__main__':
    main()